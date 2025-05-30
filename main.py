import sys
import os
import json
import datetime
import time
import asyncio
from urllib.parse import quote_plus # Để mã hóa URL an toàn

# --- Cờ Debug ---
# Đặt thành True để xem log chi tiết khi gỡ lỗi
DEBUG_WORKER = False
DEBUG_MAIN_THREAD_HANDLERS = False

from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QTextEdit, QCheckBox, QPushButton, QTableWidget, QTableWidgetItem,
    QStatusBar, QProgressBar, QSizePolicy, QHeaderView, QMessageBox,
    QFileDialog, QShortcut, QMenu, QAction, QLabel
)
from PyQt5.QtWebEngineWidgets import QWebEngineView, QWebEnginePage, QWebEngineProfile, QWebEngineSettings
from PyQt5.QtCore import Qt, QUrl, QTimer, QThread, pyqtSignal, QObject
from PyQt5.QtGui import QKeySequence, QPixmap

try:
    import openpyxl
    from openpyxl.styles import Font
except ImportError:
    # Hiển thị thông báo lỗi nếu thư viện openpyxl chưa được cài đặt
    QMessageBox.critical(None, "Lỗi thư viện",
                         "Thư viện openpyxl là bắt buộc nhưng không tìm thấy. "
                         "Vui lòng cài đặt (pip install openpyxl) và thử lại.")
    sys.exit(1)


class SearchResult:
    """Lớp lưu trữ kết quả tìm kiếm cho một từ khóa trên một website."""
    def __init__(self, website="", keyword="", link="", found="Không", full_data="", full_data_summary=""):
        self.website = website
        self.keyword = keyword
        self.link = link
        self.found = found # Trạng thái: "Có", "Không", "Lỗi", "Bị hủy"
        self.full_data = full_data # Toàn bộ text trích xuất được (nếu có)
        self.full_data_summary = full_data_summary # Tóm tắt của full_data


class SearchEngines:
    """Lớp chứa các phương thức tĩnh để tạo URL tìm kiếm cho các nền tảng."""
    @staticmethod
    def get_instagram_search_url(keyword):
        # URL cho trang tag của Instagram
        return f"https://www.instagram.com/explore/tags/{quote_plus(keyword.replace(' ', ''))}/"

    @staticmethod
    def get_facebook_top_search_url(keyword):
        return f"https://www.facebook.com/search/top/?q={quote_plus(keyword)}"

    @staticmethod
    def get_facebook_videos_search_url(keyword):
        return f"https://www.facebook.com/search/videos/?q={quote_plus(keyword)}"

    @staticmethod
    def get_facebook_posts_search_url(keyword):
        return f"https://www.facebook.com/search/posts/?q={quote_plus(keyword)}"

    @staticmethod
    def get_nike_search_url(keyword):
        return f"https://www.nike.com/vn/search?q={quote_plus(keyword)}"

    @staticmethod
    def get_x_top_search_url(keyword): # X (Twitter)
        return f"https://x.com/search?q={quote_plus(keyword)}&src=typed_query"

    @staticmethod
    def get_x_media_search_url(keyword): # X (Twitter) - Media
        return f"https://x.com/search?q={quote_plus(keyword)}&src=typed_query&f=media"

    @staticmethod
    def get_sneaker_news_search_url(keyword):
        return f"https://sneakernews.com/?s={quote_plus(keyword)}"


class SearchWorker(QObject):
    """
    Worker chạy trong một QThread riêng để thực hiện các tác vụ tìm kiếm
    mà không làm đóng băng giao diện người dùng chính.
    Sử dụng asyncio để quản lý các hoạt động I/O bất đồng bộ (tải trang, chạy JS).
    """
    finished_task = pyqtSignal(str, str, str, str, str) # website, keyword, link, found_status, full_data
    progress_update = pyqtSignal(int, str) # percent, status_message
    search_completed_all = pyqtSignal(str) # final_message

    def __init__(self, tasks_to_run, parent_app_ref):
        super().__init__()
        self.tasks_to_run = tasks_to_run # Danh sách các tác vụ (keyword, platform)
        self.parent_app = parent_app_ref # Tham chiếu đến KeywordSearchApp để truy cập config và phát tín hiệu
        self._is_running = True # Cờ để kiểm soát việc dừng worker
        self.async_loop = None # Vòng lặp sự kiện asyncio cho worker này
        self.active_async_tasks = [] # Danh sách các tác vụ asyncio đang chạy

    def stop(self):
        """Yêu cầu dừng worker và các tác vụ asyncio của nó."""
        if DEBUG_WORKER: print("WORKER: Received stop request.")
        self._is_running = False
        if self.async_loop and self.async_loop.is_running():
            if DEBUG_WORKER: print("WORKER: Cancelling active asyncio tasks...")
            for task in list(self.active_async_tasks): # Sao chép list để có thể xóa trong lúc duyệt
                if task and not task.done():
                    task.cancel()
            # Không gọi loop.stop() ở đây, để run_until_complete tự kết thúc khi các task bị hủy

    async def _safe_navigate_and_wait(self, url_string, timeout_sec=35):
        """
        Điều hướng QWebEngineView đến URL và chờ hoàn thành hoặc timeout.
        Sử dụng asyncio.Future để chờ kết quả từ luồng chính.
        """
        if not self._is_running:
            raise asyncio.CancelledError("Worker stopped during navigation.")

        qurl = QUrl(url_string)
        future = asyncio.Future(loop=self.async_loop)
        self.active_async_tasks.append(asyncio.current_task(self.async_loop))

        # Callback được gọi bởi luồng chính khi trang tải xong (hoặc lỗi)
        def callback_from_main_thread(success):
            if not future.done():
                if success:
                    future.set_result(True)
                else:
                    future.set_exception(ConnectionError(f"Navigation failed for {url_string}"))
        
        if DEBUG_WORKER: print(f"WORKER: Requesting navigation to: {url_string[:100]}")
        self.parent_app.request_load_url_from_worker.emit(qurl, callback_from_main_thread)
        
        try:
            await asyncio.wait_for(future, timeout=timeout_sec)
        except asyncio.TimeoutError:
            raise TimeoutError(f"Navigation timed out after {timeout_sec}s for {url_string}")
        except asyncio.CancelledError:
            if DEBUG_WORKER: print(f"WORKER: Navigation cancelled for {url_string[:100]}")
            # Đảm bảo future được giải quyết nếu bị hủy ở đây
            if not future.done(): future.cancel()
            raise
        finally:
            if asyncio.current_task(self.async_loop) in self.active_async_tasks:
                self.active_async_tasks.remove(asyncio.current_task(self.async_loop))


    async def _safe_execute_script_and_wait(self, script, timeout_sec=20):
        """
        Thực thi JavaScript trên QWebEngineView và chờ kết quả hoặc timeout.
        Sử dụng asyncio.Future để chờ kết quả từ luồng chính.
        """
        if not self._is_running:
            raise asyncio.CancelledError("Worker stopped during script execution.")

        future = asyncio.Future(loop=self.async_loop)
        self.active_async_tasks.append(asyncio.current_task(self.async_loop))

        # Callback được gọi bởi luồng chính khi JS thực thi xong
        def callback_from_main_thread(result):
            if not future.done():
                future.set_result(result)

        if DEBUG_WORKER: print(f"WORKER: Requesting JS execution: {script[:100]}...")
        self.parent_app.request_execute_js_from_worker.emit(script, callback_from_main_thread)

        try:
            result = await asyncio.wait_for(future, timeout=timeout_sec)
            return result
        except asyncio.TimeoutError:
            raise TimeoutError(f"JavaScript execution timed out after {timeout_sec}s.")
        except asyncio.CancelledError:
            if DEBUG_WORKER: print(f"WORKER: JS execution cancelled: {script[:100]}...")
            if not future.done(): future.cancel()
            raise
        finally:
            if asyncio.current_task(self.async_loop) in self.active_async_tasks:
                self.active_async_tasks.remove(asyncio.current_task(self.async_loop))

    def _clean_js_result(self, js_result):
        """Làm sạch kết quả trả về từ JavaScript (loại bỏ dấu ngoặc kép thừa, giải mã escape sequences)."""
        if js_result is None: return ""
        if isinstance(js_result, str):
            # Loại bỏ dấu "" ở đầu và cuối nếu có (thường do JSON.stringify)
            if js_result.startswith("\"") and js_result.endswith("\"") and len(js_result) > 1:
                js_result = js_result[1:-1]
            # Thay thế các escape sequences phổ biến
            js_result = (js_result.replace("\\n", "\n")
                         .replace("\\r", "\r")
                         .replace("\\t", "\t")
                         .replace("\\\"", "\"")
                         .replace("\\\\", "\\"))
        return js_result

    async def _perform_instagram_search(self, keyword):
        """Thực hiện tìm kiếm trên Instagram cho một từ khóa cụ thể."""
        if not self._is_running: return
        self.progress_update.emit(0, f"Instagram: Bắt đầu tìm '{keyword}'...")
        base_tag_url = SearchEngines.get_instagram_search_url(keyword)
        found_links_for_keyword = set() # Để tránh xử lý trùng lặp link nếu có
        max_posts_to_process = self.parent_app.config_insta_max_posts # Giới hạn số bài đăng xử lý

        try:
            await self._safe_navigate_and_wait(base_tag_url)
            await asyncio.sleep(self.parent_app.config_wait_short) # Chờ trang tag tải

            # Cuộn xuống vài lần để tải thêm bài đăng
            for i in range(self.parent_app.config_insta_scrolls):
                if not self._is_running: return
                self.progress_update.emit(int((i+1) * 5 / self.parent_app.config_insta_scrolls), f"Instagram: Đang cuộn trang tag ({i+1})...")
                await self._safe_execute_script_and_wait("window.scrollTo(0, document.body.scrollHeight);")
                await asyncio.sleep(self.parent_app.config_wait_medium)

            # JavaScript để lấy link các bài đăng trên trang tag
            # Selector này rất dễ thay đổi, cần cập nhật thường xuyên
            js_get_post_links = """
                (function(){
                    var links = new Set();
                    // Selector cho các bài đăng (hình ảnh/video/reels)
                    // Thử các selector phổ biến, từ cụ thể đến chung chung
                    var selectors = [
                        'main article a[href*="/p/"]', 'main article a[href*="/reel/"]', // Cấu trúc cũ hơn
                        'div[role="main"] a[href*="/p/"]', 'div[role="main"] a[href*="/reel/"]', // Cấu trúc có role="main"
                        'main div section a[href*="/p/"]', 'main div section a[href*="/reel/"]',
                        'a[href*="/p/"]', 'a[href*="/reel/"]' // Chung nhất
                    ];
                    for (var s_idx = 0; s_idx < selectors.length; s_idx++) {
                        var nodes = document.querySelectorAll(selectors[s_idx]);
                        for(var i = 0; i < nodes.length; i++){
                            if(nodes[i].href) { links.add(nodes[i].href); }
                        }
                        if (links.size > 0 && s_idx < selectors.length -1) break; // Nếu tìm thấy bằng selector cụ thể thì dừng
                    }
                    // Giới hạn số lượng link (ví dụ: 20 link đầu tiên)
                    return JSON.stringify(Array.from(links).slice(0, 20));
                })();"""
            
            post_links_json = await self._safe_execute_script_and_wait(js_get_post_links)
            post_links_raw = json.loads(self._clean_js_result(post_links_json) or "[]")
            
            if not post_links_raw:
                self.finished_task.emit("Instagram", keyword, base_tag_url, "Không (no posts)", "Không tìm thấy bài đăng nào trên trang tag.")
                return

            self.progress_update.emit(10, f"Instagram: Tìm thấy {len(post_links_raw)} link bài đăng tiềm năng cho '{keyword}'.")
            
            item_processed_count = 0
            for post_url in post_links_raw:
                if not self._is_running: return
                if item_processed_count >= max_posts_to_process:
                    if DEBUG_WORKER: print(f"WORKER: Instagram reached max posts ({max_posts_to_process}) for '{keyword}'.")
                    break
                if post_url in found_links_for_keyword: continue # Bỏ qua nếu link này đã được xử lý
                
                item_processed_count +=1
                self.progress_update.emit(10 + int(item_processed_count * 80 / min(len(post_links_raw), max_posts_to_process)),
                                          f"Instagram: Đang xử lý bài đăng {item_processed_count}/{min(len(post_links_raw), max_posts_to_process)} cho '{keyword}'...")
                try:
                    await self._safe_navigate_and_wait(post_url)
                    await asyncio.sleep(self.parent_app.config_wait_short)

                    # JS để lấy caption/text từ trang chi tiết bài đăng
                    # Selector này cũng rất dễ thay đổi
                    js_get_caption = """
                        (function() {
                            let caption = "";
                            // Thử selector cho caption chính (thường là H1 hoặc phần tử có testid)
                            let mainCaption = document.querySelector('h1, [data-testid="post-caption"]');
                            if (mainCaption && mainCaption.innerText.trim()) {
                                caption = mainCaption.innerText.trim();
                            } else {
                                // Thử selector cho các bình luận đầu tiên / mô tả có thể chứa caption
                                let commentNodes = document.querySelectorAll('article div[role="button"] + div ul li span, article div[role="dialog"] div[role="dialog"] ul li span, div[data-testid="caption"] span');
                                if (commentNodes.length > 0 && commentNodes[0].innerText.trim()) {
                                     caption = commentNodes[0].innerText.trim();
                                } else {
                                    // Selector khác cho trường hợp caption nằm trong một cấu trúc khác (thường là span có dir="auto")
                                    let elementsWithDirAuto = document.querySelectorAll('span[dir="auto"], div[dir="auto"]');
                                    for (let el of elementsWithDirAuto) {
                                        // Kiểm tra xem phần tử có phải là con của một link không (để tránh lấy tên người dùng)
                                        // và có text đủ dài
                                        if (el.innerText && el.innerText.length > 20 && !el.closest('a')) {
                                            caption = el.innerText.trim();
                                            break;
                                        }
                                    }
                                }
                            }
                            // Fallback: lấy text từ phần header của bài viết nếu có
                            if (!caption) {
                                let headerText = document.querySelector('article header');
                                if (headerText) caption = headerText.innerText.trim();
                            }
                            // Fallback cuối cùng: toàn bộ body (ít chính xác hơn, chỉ lấy một phần)
                            if (!caption) caption = document.body.innerText.substring(0, 2000);

                            return caption;
                        })();
                    """
                    post_text_raw = await self._safe_execute_script_and_wait(js_get_caption)
                    post_text = self._clean_js_result(post_text_raw)

                    found_status = "Có" if keyword.lower() in post_text.lower() else "Không"
                    self.finished_task.emit("Instagram", keyword, post_url, found_status, post_text)
                    found_links_for_keyword.add(post_url)

                    if found_status == "Có" and self.parent_app.cb_long_screenshot.isChecked():
                        self.parent_app.request_scroll_to_keyword_from_worker.emit(keyword)
                        await asyncio.sleep(1.5) # Chờ scroll
                        self.parent_app.request_capture_screenshot_from_worker.emit("Instagram", keyword, post_url)
                        await asyncio.sleep(1) # Chờ screenshot
                
                except TimeoutError as te_post:
                    self.finished_task.emit("Instagram", keyword, post_url, "Lỗi (Timeout Post)", str(te_post))
                except ConnectionError as ce_post:
                    self.finished_task.emit("Instagram", keyword, post_url, "Lỗi (Nav Post)", str(ce_post))
                except asyncio.CancelledError:
                    self.finished_task.emit("Instagram", keyword, post_url, "Bị hủy", "Tác vụ bài đăng bị hủy.")
                    return
                except Exception as e_post:
                    self.finished_task.emit("Instagram", keyword, post_url, "Lỗi (Post)", f"{type(e_post).__name__}: {str(e_post)}")
                
                await asyncio.sleep(self.parent_app.config_wait_short) # Nghỉ giữa các bài đăng

            if item_processed_count == 0 and len(post_links_raw) > 0:
                 self.finished_task.emit("Instagram", keyword, base_tag_url, "Không (processing error)", "Có link nhưng không xử lý được bài đăng nào.")

        except TimeoutError as te:
            self.finished_task.emit("Instagram", keyword, base_tag_url, "Lỗi (Timeout)", str(te))
        except ConnectionError as ce:
            self.finished_task.emit("Instagram", keyword, base_tag_url, "Lỗi (Navigation)", str(ce))
        except asyncio.CancelledError:
            self.finished_task.emit("Instagram", keyword, base_tag_url, "Bị hủy", "Tác vụ Instagram bị hủy.")
        except Exception as e:
            self.finished_task.emit("Instagram", keyword, base_tag_url, "Lỗi", f"{type(e).__name__}: {str(e)}")

    async def _perform_generic_search(self, url_generator, website_name, keyword):
        """Thực hiện tìm kiếm chung cho các nền tảng khác (Facebook, Nike, X, SneakerNews)."""
        if not self._is_running: return
        self.progress_update.emit(0, f"{website_name}: Bắt đầu tìm '{keyword}'...")
        search_url = url_generator(keyword)
        try:
            await self._safe_navigate_and_wait(search_url)
            await asyncio.sleep(self.parent_app.config_wait_medium) # Chờ trang tải và render

            # Xóa các tiêu đề "Kết quả tìm kiếm cho..." để không ảnh hưởng đến việc tìm từ khóa
            if "Facebook" in website_name or "Nike" in website_name:
                js_remove_h1 = """
                    (function() {
                        let h1s = document.querySelectorAll('h1');
                        h1s.forEach(h1 => {
                            if (h1.innerText.toLowerCase().includes('kết quả tìm kiếm cho') || 
                                h1.innerText.toLowerCase().includes('search results for')) {
                                h1.remove();
                            }
                        });
                    })();"""
                await self._safe_execute_script_and_wait(js_remove_h1, timeout_sec=5) # Timeout ngắn hơn cho script này

            scroll_count = 0
            if "Facebook" in website_name or "X" in website_name:
                scroll_count = self.parent_app.config_fb_x_scrolls
            elif "Nike" in website_name or "Sneaker News" in website_name:
                scroll_count = 1 # Các trang này có thể không cần cuộn nhiều hoặc có pagination
            
            for i in range(scroll_count):
                if not self._is_running: return
                self.progress_update.emit(int((i+1) * 20 / scroll_count) if scroll_count > 0 else 0,
                                          f"{website_name}: Đang cuộn trang ({i+1}/{scroll_count})...")
                await self._safe_execute_script_and_wait(f"window.scrollBy(0, window.innerHeight * {0.8 + i*0.2});")
                await asyncio.sleep(self.parent_app.config_wait_short)

            js_get_data = "document.body.innerText;"
            full_data_raw = await self._safe_execute_script_and_wait(js_get_data)
            full_data = self._clean_js_result(full_data_raw)

            found_status = "Có" if keyword.lower() in full_data.lower() else "Không"
            self.finished_task.emit(website_name, keyword, search_url, found_status, full_data)

            if found_status == "Có" and self.parent_app.cb_long_screenshot.isChecked():
                self.parent_app.request_scroll_to_keyword_from_worker.emit(keyword)
                await asyncio.sleep(1.5)
                self.parent_app.request_capture_screenshot_from_worker.emit(website_name, keyword, search_url)
                await asyncio.sleep(1)

        except TimeoutError as te:
            self.finished_task.emit(website_name, keyword, search_url, "Lỗi (Timeout)", str(te))
        except ConnectionError as ce:
            self.finished_task.emit(website_name, keyword, search_url, "Lỗi (Navigation)", str(ce))
        except asyncio.CancelledError:
            self.finished_task.emit(website_name, keyword, search_url, "Bị hủy", "Tác vụ bị hủy.")
        except Exception as e:
            self.finished_task.emit(website_name, keyword, search_url, "Lỗi", f"{type(e).__name__}: {str(e)}")

    async def _run_all_tasks_async(self):
        """Hàm chính chạy bất đồng bộ tất cả các tác vụ tìm kiếm."""
        total_tasks_overall = len(self.tasks_to_run)
        current_task_num_overall = 0

        for task_info in self.tasks_to_run:
            if not self._is_running:
                self.progress_update.emit(100, "Tìm kiếm bị dừng bởi người dùng.")
                break
            
            current_task_num_overall += 1
            # Tính toán tiến trình tổng thể một cách cẩn thận hơn
            # Giả sử mỗi task_info là một cặp (keyword, platform)
            # Nếu có N keywords và M platforms, tổng số "sub-tasks" là N*M
            # tasks_to_run đã được tạo ra với mỗi cặp (keyword, platform) là một item
            progress_percent = int(current_task_num_overall * 100 / total_tasks_overall) if total_tasks_overall > 0 else 0
            
            keyword = task_info["keyword"]
            platform_name = task_info["platform_name"]
            url_generator = task_info.get("url_generator") # Có thể None cho Instagram

            self.progress_update.emit(progress_percent,
                                      f"Đang xử lý {platform_name} cho '{keyword}' ({current_task_num_overall}/{total_tasks_overall})")
            try:
                current_async_task = None
                if platform_name == "Instagram":
                    current_async_task = self.async_loop.create_task(self._perform_instagram_search(keyword))
                else:
                    if url_generator:
                        current_async_task = self.async_loop.create_task(self._perform_generic_search(url_generator, platform_name, keyword))
                    else:
                        self.finished_task.emit(platform_name, keyword, "N/A", "Lỗi (Config)", "URL generator không được định nghĩa")
                        continue # Bỏ qua task này
                
                if current_async_task:
                    self.active_async_tasks.append(current_async_task)
                    await current_async_task # Chờ tác vụ con hoàn thành
                    if current_async_task in self.active_async_tasks: # Xóa sau khi hoàn thành
                         self.active_async_tasks.remove(current_async_task)

            except asyncio.CancelledError:
                if DEBUG_WORKER: print(f"WORKER: Task for {platform_name} - '{keyword}' was cancelled.")
                self.progress_update.emit(progress_percent, f"Tác vụ cho {platform_name} - '{keyword}' đã bị hủy.")
                # Không break ở đây, để vòng lặp chính kiểm tra _is_running
            except Exception as e_task_run:
                if DEBUG_WORKER: print(f"WORKER: Unexpected error running task for {platform_name} - '{keyword}': {e_task_run}")
                self.finished_task.emit(platform_name, keyword, "N/A", "Lỗi (Worker)", f"Lỗi không xác định: {e_task_run}")
            
            if not self._is_running: # Kiểm tra lại sau mỗi task lớn
                self.progress_update.emit(100, "Tìm kiếm bị dừng giữa chừng.")
                break
            
            await asyncio.sleep(0.2) # Một khoảng nghỉ nhỏ giữa các tác vụ lớn (keyword-platform)

        if self._is_running:
            self.search_completed_all.emit(f"Hoàn tất tìm kiếm. Đã xử lý {current_task_num_overall} tác vụ.")
        else:
            self.search_completed_all.emit(f"Tìm kiếm đã dừng. Đã xử lý {current_task_num_overall} tác vụ.")

    def run(self):
        """Phương thức chính được QThread gọi khi bắt đầu."""
        if DEBUG_WORKER: print("WORKER: Thread started, setting up asyncio loop.")
        self.async_loop = asyncio.new_event_loop()
        asyncio.set_event_loop(self.async_loop)
        
        try:
            # Chạy hàm _run_all_tasks_async trong vòng lặp asyncio
            self.async_loop.run_until_complete(self._run_all_tasks_async())
        except Exception as e_loop:
            if DEBUG_WORKER: print(f"WORKER: Critical error in worker's asyncio loop: {e_loop}")
            self.search_completed_all.emit(f"Lỗi nghiêm trọng trong luồng worker: {e_loop}")
        finally:
            if DEBUG_WORKER: print("WORKER: Asyncio loop completing. Cleaning up...")
            try:
                # Dọn dẹp các tác vụ asyncio còn lại nếu vòng lặp kết thúc đột ngột hoặc bị hủy
                if self.async_loop.is_running(): # Kiểm tra xem loop có đang chạy không
                    pending_tasks = [task for task in asyncio.all_tasks(self.async_loop) if task is not asyncio.current_task(self.async_loop)]
                    if pending_tasks:
                        if DEBUG_WORKER: print(f"WORKER: Found {len(pending_tasks)} pending tasks to cancel during cleanup.")
                        for task in pending_tasks:
                            if not task.done():
                                task.cancel()
                        # Cho các tác vụ một chút thời gian để hủy
                        self.async_loop.run_until_complete(asyncio.gather(*pending_tasks, return_exceptions=True))
            except Exception as e_cleanup_tasks:
                if DEBUG_WORKER: print(f"WORKER: Error during asyncio task cleanup: {e_cleanup_tasks}")
            
            if self.async_loop:
                self.async_loop.close() # Đóng vòng lặp asyncio
            self.async_loop = None
            self.active_async_tasks.clear()
            if DEBUG_WORKER: print("WORKER: Thread run method finished and loop closed.")


class KeywordSearchApp(QMainWindow):
    """Lớp chính của ứng dụng, quản lý giao diện và luồng tìm kiếm."""
    # Tín hiệu để yêu cầu QWebEngineView (chạy trên luồng chính) thực hiện hành động từ worker
    request_load_url_from_worker = pyqtSignal(QUrl, object) # URL, callback_cho_future_cua_worker
    request_execute_js_from_worker = pyqtSignal(str, object) # script, callback_cho_future_cua_worker
    request_scroll_to_keyword_from_worker = pyqtSignal(str) # keyword
    request_capture_screenshot_from_worker = pyqtSignal(str, str, str) # website, keyword, identifier (URL/timestamp)

    def __init__(self):
        super().__init__()
        self.setWindowTitle("Keyword Search Tool - Python/PyQt5 (v3)")
        self.setGeometry(50, 50, 1400, 950) # Tăng kích thước cửa sổ một chút

        # --- Cấu hình ---
        self.config_wait_short = 2.8  # Thời gian chờ ngắn (ví dụ: sau khi cuộn, tải bài đăng)
        self.config_wait_medium = 4.5 # Thời gian chờ trung bình (ví dụ: tải trang tag)
        # self.config_wait_long = 7.0 # Hiện không dùng nhiều
        self.config_insta_scrolls = 2      # Số lần cuộn trên trang tag Instagram
        self.config_insta_max_posts = 10   # Số bài đăng Instagram tối đa xử lý cho mỗi từ khóa
        self.config_fb_x_scrolls = 2       # Số lần cuộn trên trang Facebook/X

        self.results_list = [] # Danh sách lưu trữ các đối tượng SearchResult
        self.search_worker_obj = None # Đối tượng worker
        self.search_thread = None # QThread cho worker
        self.stopwatch_start_time = None # Thời điểm bắt đầu tìm kiếm
        self.elapsed_timer_display = QTimer(self) # Timer để cập nhật thời gian trôi qua
        self.elapsed_timer_display.setInterval(1000) # Cập nhật mỗi giây
        self.current_search_tasks_list = [] # Danh sách các tác vụ tìm kiếm hiện tại
        self.active_web_callbacks = {} # Dictionary để quản lý các callback động cho QWebEngineView

        self._init_ui()
        self._setup_connections()
        self._setup_shortcuts()
        self.load_keywords_from_file() # Tải từ khóa từ file khi khởi động
        self.update_status("Ứng dụng đã khởi tạo. Sẵn sàng tìm kiếm.")

    def _init_ui(self):
        """Khởi tạo các thành phần giao diện người dùng."""
        main_widget = QWidget()
        self.setCentralWidget(main_widget)
        main_layout = QHBoxLayout(main_widget)

        # --- Bảng điều khiển bên trái ---
        left_panel = QWidget()
        left_layout = QVBoxLayout(left_panel)
        left_panel.setMinimumWidth(320)
        left_panel.setMaximumWidth(400)

        self.txt_keyword = QTextEdit()
        self.txt_keyword.setPlaceholderText("Nhập từ khóa, mỗi từ khóa trên một dòng...")
        left_layout.addWidget(self.txt_keyword)

        # Checkbox chọn nền tảng
        self.cb_instagram = QCheckBox("Instagram")
        self.cb_facebook_top = QCheckBox("Facebook Top")
        self.cb_facebook_videos = QCheckBox("Facebook Videos")
        self.cb_facebook_posts = QCheckBox("Facebook Posts")
        self.cb_nike = QCheckBox("Nike")
        self.cb_x_top = QCheckBox("X Top (Twitter)")
        self.cb_x_media = QCheckBox("X Media (Twitter)")
        self.cb_sneaker_news = QCheckBox("Sneaker News")

        self.cb_instagram.setChecked(True) # Mặc định chọn Instagram

        search_options_layout = QVBoxLayout()
        checkboxes = [self.cb_instagram, self.cb_facebook_top, self.cb_facebook_videos,
                      self.cb_facebook_posts, self.cb_nike, self.cb_x_top,
                      self.cb_x_media, self.cb_sneaker_news]
        for cb in checkboxes: search_options_layout.addWidget(cb)
        left_layout.addLayout(search_options_layout)

        # Các nút điều khiển
        self.btn_search = QPushButton("Tìm kiếm (Ctrl+S)")
        self.btn_search.setStyleSheet("background-color: #4CAF50; color: white; padding: 8px;")
        left_layout.addWidget(self.btn_search)
        
        self.btn_stop_search = QPushButton("Dừng Tìm kiếm (Esc)")
        self.btn_stop_search.setStyleSheet("background-color: #f44336; color: white; padding: 8px;")
        self.btn_stop_search.setEnabled(False)
        left_layout.addWidget(self.btn_stop_search)

        self.btn_login = QPushButton("Đăng nhập vào Nền tảng")
        left_layout.addWidget(self.btn_login)
        self.btn_export_excel = QPushButton("Xuất ra Excel (Ctrl+E)")
        left_layout.addWidget(self.btn_export_excel)
        self.btn_update_keywords = QPushButton("Cập nhật File Từ khóa (Ctrl+U)")
        left_layout.addWidget(self.btn_update_keywords)
        self.btn_clear_cache = QPushButton("Xóa Cache Trình duyệt")
        left_layout.addWidget(self.btn_clear_cache)

        # Tùy chọn thêm
        self.cb_auto_xlsx = QCheckBox("Tự động lưu Excel sau khi tìm")
        self.cb_auto_xlsx.setChecked(True)
        left_layout.addWidget(self.cb_auto_xlsx)
        
        self.cb_long_screenshot = QCheckBox("Thử Chụp Ảnh Dài (nếu tìm thấy)")
        self.cb_long_screenshot.setToolTip("Tính năng này chụp toàn bộ trang, có thể không ổn định và tốn thời gian.")
        left_layout.addWidget(self.cb_long_screenshot)

        left_layout.addStretch(1) # Đẩy các widget lên trên
        main_layout.addWidget(left_panel)

        # --- Bảng điều khiển bên phải ---
        right_panel = QWidget()
        right_layout = QVBoxLayout(right_panel)
        
        self.web_view = QWebEngineView()
        # Cấu hình QWebEngineView
        self.web_view.settings().setAttribute(QWebEngineSettings.JavascriptEnabled, True)
        self.web_view.settings().setAttribute(QWebEngineSettings.ScrollAnimatorEnabled, True) # Cho phép cuộn mượt
        self.web_view.settings().setAttribute(QWebEngineSettings.FullScreenSupportEnabled, True)
        self.web_view.settings().setAttribute(QWebEngineSettings.AllowRunningInsecureContent, False) # Tăng bảo mật
        # Cập nhật User-Agent để giống trình duyệt hiện đại hơn
        user_agent = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/110.0.0.0 Safari/537.36 Edg/110.0.1587.41"
        self.web_view.page().profile().setHttpUserAgent(user_agent)
        self.web_view.load(QUrl("https://www.google.com.vn/?hl=vi")) # Trang chủ mặc định
        right_layout.addWidget(self.web_view, 3) # Webview chiếm 3 phần

        self.dgv_results = QTableWidget() # Bảng hiển thị kết quả
        self.dgv_results.setColumnCount(5)
        self.dgv_results.setHorizontalHeaderLabels(["Website", "Keyword", "Link", "Found", "Data Summary"])
        header = self.dgv_results.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(1, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(2, QHeaderView.Stretch) # Cột link co giãn
        header.setSectionResizeMode(3, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(4, QHeaderView.Interactive) # Cho phép người dùng thay đổi kích thước cột summary
        self.dgv_results.setColumnWidth(4, 350) # Tăng chiều rộng mặc định cho cột summary
        self.dgv_results.setEditTriggers(QTableWidget.NoEditTriggers) # Không cho sửa trực tiếp trên bảng
        self.dgv_results.setSelectionBehavior(QTableWidget.SelectRows) # Chọn cả hàng
        self.dgv_results.setWordWrap(True) # Tự động xuống dòng trong ô
        self.dgv_results.setStyleSheet("QTableWidget { alternate-background-color: #f0f0f0; }"
                                       "QHeaderView::section { background-color: #e0e0e0; padding: 4px; border: 1px solid #d0d0d0; }")
        right_layout.addWidget(self.dgv_results, 2) # Bảng kết quả chiếm 2 phần

        main_layout.addWidget(right_panel, 1) # right_panel chiếm tỷ lệ lớn hơn

        # --- Thanh trạng thái ---
        self.status_bar = QStatusBar()
        self.setStatusBar(self.status_bar)
        self.lbl_status_main_text = QLabel("Sẵn sàng")
        self.status_bar.addWidget(self.lbl_status_main_text, 1) # Chiếm phần lớn thanh trạng thái

        self.lbl_elapsed_time_display_widget = QLabel("Thời gian: 00:00:00")
        self.status_bar.addPermanentWidget(self.lbl_elapsed_time_display_widget) # Hiển thị cố định bên phải

        self.progress_bar = QProgressBar()
        self.status_bar.addPermanentWidget(self.progress_bar)
        self.progress_bar.setVisible(False) # Ban đầu ẩn
        self.progress_bar.setMaximumWidth(280)
        self.progress_bar.setTextVisible(False)

    def _setup_connections(self):
        """Kết nối các tín hiệu (signals) và khe cắm (slots)."""
        # Kết nối từ Main App (GUI thread) đến Worker (chạy trên QThread)
        # được thực hiện khi tạo worker (ví dụ: self.search_worker_obj.moveToThread)

        # Kết nối các yêu cầu từ Worker (chạy trên QThread) đến các handler trên Main App (GUI thread)
        self.request_load_url_from_worker.connect(self.handle_load_url_request_on_main)
        self.request_execute_js_from_worker.connect(self.handle_execute_js_request_on_main)
        self.request_scroll_to_keyword_from_worker.connect(self.scroll_to_keyword_on_main)
        self.request_capture_screenshot_from_worker.connect(self.capture_screenshot_on_main)

        # Kết nối các button UI
        self.btn_search.clicked.connect(self.start_search)
        self.btn_stop_search.clicked.connect(self.stop_current_search)
        self.btn_login.clicked.connect(self.show_login_menu)
        self.btn_export_excel.clicked.connect(self.export_to_excel)
        self.btn_update_keywords.clicked.connect(self.update_keywords_file)
        self.btn_clear_cache.clicked.connect(self.clear_browser_cache)

        # Kết nối timer cập nhật thời gian
        self.elapsed_timer_display.timeout.connect(self.update_elapsed_time_display_label)
        
        # Kết nối sự kiện click vào link trên bảng kết quả
        self.dgv_results.cellDoubleClicked.connect(self.open_link_from_table)


    def _setup_shortcuts(self):
        """Thiết lập các phím tắt cho ứng dụng."""
        QShortcut(QKeySequence("Ctrl+S"), self, self.start_search)
        QShortcut(QKeySequence("Ctrl+E"), self, self.export_to_excel)
        QShortcut(QKeySequence("Ctrl+U"), self, self.update_keywords_file)
        QShortcut(QKeySequence("F5"), self, lambda: self.web_view.reload()) # F5 để tải lại trang web
        QShortcut(QKeySequence(Qt.Key_Escape), self, self.stop_current_search_if_running) # Esc để dừng tìm kiếm


    def handle_load_url_request_on_main(self, qurl, callback_for_future):
        """Xử lý yêu cầu tải URL từ worker, chạy trên luồng chính."""
        if DEBUG_MAIN_THREAD_HANDLERS: print(f"MAIN: Received load URL request for: {qurl.toString()[:100]}")
        # Tạo một ID duy nhất cho callback này để quản lý
        # Điều này quan trọng nếu có nhiều yêu cầu đồng thời (mặc dù worker hiện tại xử lý tuần tự)
        callback_id = id(callback_for_future) 
        self.active_web_callbacks[callback_id] = callback_for_future
        
        # Ngắt kết nối cũ nếu có để tránh gọi nhiều lần cho cùng một sự kiện loadFinished
        try:
            self.web_view.page().loadFinished.disconnect()
        except TypeError: # Nếu chưa có kết nối nào trước đó
            pass
        
        # Kết nối loadFinished với một lambda chứa callback_id
        self.web_view.page().loadFinished.connect(
            lambda success: self._on_main_load_finished(success, callback_id)
        )
        self.web_view.load(qurl)

    def _on_main_load_finished(self, success, callback_id):
        """Được gọi khi QWebEnginePage tải xong, chạy trên luồng chính."""
        if DEBUG_MAIN_THREAD_HANDLERS: print(f"MAIN: Load finished (success: {success}) for callback_id: {callback_id}")
        if callback_id in self.active_web_callbacks:
            callback_func = self.active_web_callbacks.pop(callback_id) # Lấy và xóa callback
            if callback_func:
                try:
                    callback_func(success) # Gọi callback đã được truyền từ worker
                except Exception as e_cb:
                    if DEBUG_MAIN_THREAD_HANDLERS: print(f"MAIN: Error in load_finished callback for worker: {e_cb}")
        # Ngắt kết nối sau khi sử dụng để tránh gọi lại cho các lần load khác không liên quan
        try:
            self.web_view.page().loadFinished.disconnect()
        except TypeError:
            pass

    def handle_execute_js_request_on_main(self, script, callback_for_future):
        """Xử lý yêu cầu thực thi JavaScript từ worker, chạy trên luồng chính."""
        if DEBUG_MAIN_THREAD_HANDLERS: print(f"MAIN: Received JS execution request: {script[:100]}...")
        callback_id = id(callback_for_future)
        self.active_web_callbacks[callback_id] = callback_for_future
        
        # Hàm callback nội bộ để gọi callback của worker
        def internal_js_callback_on_main(result):
            if DEBUG_MAIN_THREAD_HANDLERS: print(f"MAIN: JS result received for callback_id: {callback_id}")
            if callback_id in self.active_web_callbacks:
                worker_callback = self.active_web_callbacks.pop(callback_id)
                if worker_callback:
                    try:
                        worker_callback(result)
                    except Exception as e_cb_js:
                        if DEBUG_MAIN_THREAD_HANDLERS: print(f"MAIN: Error in JS callback for worker: {e_cb_js}")
        
        self.web_view.page().runJavaScript(script, internal_js_callback_on_main)

    def update_status(self, message):
        """Cập nhật thông báo trên thanh trạng thái."""
        self.lbl_status_main_text.setText(message)
        if DEBUG_MAIN_THREAD_HANDLERS: print(f"STATUS: {message}") # In ra console nếu debug

    def update_progress_bar(self, value, status_text=""):
        """Cập nhật thanh tiến trình và thông báo trạng thái."""
        if not self.progress_bar.isVisible():
            self.progress_bar.setVisible(True)
        self.progress_bar.setValue(value)
        if status_text:
            self.update_status(status_text)

    def update_elapsed_time_display_label(self):
        """Cập nhật hiển thị thời gian đã trôi qua."""
        if self.stopwatch_start_time is not None:
            elapsed_seconds = int(time.time() - self.stopwatch_start_time)
            h = elapsed_seconds // 3600
            m = (elapsed_seconds % 3600) // 60
            s = elapsed_seconds % 60
            self.lbl_elapsed_time_display_widget.setText(f"Thời gian: {h:02}:{m:02}:{s:02}")
        else:
            self.lbl_elapsed_time_display_widget.setText("Thời gian: 00:00:00")

    def load_keywords_from_file(self):
        """Tải danh sách từ khóa từ file keywords.txt."""
        try:
            file_path = "keywords.txt" # File nằm cùng thư mục với script
            if os.path.exists(file_path):
                with open(file_path, "r", encoding="utf-8") as f:
                    self.txt_keyword.setPlainText(f.read())
                self.update_status("Đã tải từ khóa từ keywords.txt")
            else:
                self.update_status("Không tìm thấy file keywords.txt. Vui lòng nhập từ khóa thủ công.")
        except Exception as e:
            QMessageBox.warning(self, "Lỗi Đọc File", f"Không thể tải từ khóa từ file: {e}")

    def update_keywords_file(self):
        """Lưu/Cập nhật nội dung từ ô nhập liệu vào file keywords.txt."""
        try:
            file_path = "keywords.txt"
            with open(file_path, "w", encoding="utf-8") as f:
                f.write(self.txt_keyword.toPlainText())
            self.update_status("Đã cập nhật và lưu từ khóa vào keywords.txt.")
            QMessageBox.information(self, "Đã lưu Từ khóa", "Đã cập nhật và lưu từ khóa vào file keywords.txt.")
        except Exception as e:
            QMessageBox.critical(self, "Lỗi Ghi File", f"Không thể lưu từ khóa vào file: {e}")

    def build_search_tasks_list(self, keywords_list_str):
        """Xây dựng danh sách các tác vụ tìm kiếm dựa trên từ khóa và nền tảng được chọn."""
        tasks = []
        for keyword_str in keywords_list_str:
            if not keyword_str.strip(): continue # Bỏ qua dòng trống
            keyword_actual = keyword_str.strip()

            if self.cb_instagram.isChecked():
                tasks.append({"platform_name": "Instagram", "keyword": keyword_actual}) # Không cần url_generator vì có logic riêng
            if self.cb_facebook_top.isChecked():
                tasks.append({"platform_name": "Facebook Top", "keyword": keyword_actual, "url_generator": SearchEngines.get_facebook_top_search_url})
            if self.cb_facebook_videos.isChecked():
                tasks.append({"platform_name": "Facebook Videos", "keyword": keyword_actual, "url_generator": SearchEngines.get_facebook_videos_search_url})
            if self.cb_facebook_posts.isChecked():
                tasks.append({"platform_name": "Facebook Posts", "keyword": keyword_actual, "url_generator": SearchEngines.get_facebook_posts_search_url})
            if self.cb_nike.isChecked():
                tasks.append({"platform_name": "Nike", "keyword": keyword_actual, "url_generator": SearchEngines.get_nike_search_url})
            if self.cb_x_top.isChecked():
                tasks.append({"platform_name": "X Top", "keyword": keyword_actual, "url_generator": SearchEngines.get_x_top_search_url})
            if self.cb_x_media.isChecked():
                tasks.append({"platform_name": "X Media", "keyword": keyword_actual, "url_generator": SearchEngines.get_x_media_search_url})
            if self.cb_sneaker_news.isChecked():
                tasks.append({"platform_name": "Sneaker News", "keyword": keyword_actual, "url_generator": SearchEngines.get_sneaker_news_search_url})
        return tasks

    def start_search(self):
        """Bắt đầu quá trình tìm kiếm."""
        if self.search_thread and self.search_thread.isRunning():
            QMessageBox.information(self, "Đang Tìm kiếm", "Một quá trình tìm kiếm khác đang chạy. Vui lòng đợi hoặc dừng nó lại.")
            return

        keywords = [k.strip() for k in self.txt_keyword.toPlainText().splitlines() if k.strip()]
        if not keywords:
            QMessageBox.information(self, "Thiếu Từ khóa", "Vui lòng nhập ít nhất một từ khóa.")
            return

        self.current_search_tasks_list = self.build_search_tasks_list(keywords)
        if not self.current_search_tasks_list:
            QMessageBox.information(self, "Thiếu Nền tảng", "Vui lòng chọn ít nhất một nền tảng để tìm kiếm.")
            return

        self.btn_search.setEnabled(False)
        self.btn_stop_search.setEnabled(True)
        self.dgv_results.setRowCount(0) # Xóa kết quả cũ trên bảng
        self.results_list.clear() # Xóa danh sách kết quả cũ
        self.active_web_callbacks.clear() # Xóa các callbacks cũ chưa được xử lý (nếu có)

        self.stopwatch_start_time = time.time()
        self.elapsed_timer_display.start()
        self.update_status("Bắt đầu tìm kiếm...")
        self.update_progress_bar(0, "Đang khởi tạo worker...")

        # Khởi tạo worker và thread mới cho mỗi lần tìm kiếm
        self.search_worker_obj = SearchWorker(self.current_search_tasks_list, self)
        self.search_thread = QThread()
        self.search_worker_obj.moveToThread(self.search_thread) # Di chuyển worker sang thread mới

        # Kết nối tín hiệu từ worker
        self.search_worker_obj.finished_task.connect(self.add_result_to_grid_display)
        self.search_worker_obj.progress_update.connect(self.update_progress_bar)
        self.search_worker_obj.search_completed_all.connect(self.handle_search_completion_event)
        
        # Dọn dẹp luồng và worker sau khi chúng kết thúc một cách an toàn
        self.search_thread.finished.connect(self.search_thread.deleteLater) # Xóa QThread khi xong
        self.search_worker_obj.destroyed.connect(self._on_worker_destroyed) # Khi worker bị xóa

        self.search_thread.started.connect(self.search_worker_obj.run) # Gọi hàm run của worker khi thread bắt đầu
        self.search_thread.start() # Bắt đầu thread

    def _on_worker_destroyed(self):
        """Được gọi khi đối tượng worker đã bị hủy."""
        if DEBUG_MAIN_THREAD_HANDLERS: print("MAIN: Worker object has been destroyed.")
        self.search_worker_obj = None # Đặt lại tham chiếu để tránh sử dụng đối tượng đã bị xóa

    def stop_current_search_if_running(self):
        """Dừng tìm kiếm nếu đang chạy (thường gọi bằng phím Esc)."""
        if self.search_thread and self.search_thread.isRunning():
            self.stop_current_search()

    def stop_current_search(self):
        """Yêu cầu dừng quá trình tìm kiếm hiện tại."""
        self.update_status("Đang yêu cầu dừng tìm kiếm...")
        self.btn_stop_search.setEnabled(False) # Vô hiệu hóa ngay để tránh click nhiều lần
        
        if self.search_worker_obj:
            self.search_worker_obj.stop() # Gửi tín hiệu dừng cho worker
        
        # Worker sẽ tự xử lý việc dừng các tác vụ asyncio và kết thúc.
        # QThread.finished sẽ được phát ra, sau đó handle_search_completion_event và cleanup_after_search_processing sẽ được gọi.
        # Không nên gọi QThread.quit() hoặc terminate() trực tiếp ở đây vì worker đang quản lý asyncio loop.

    def handle_search_completion_event(self, final_message):
        """Xử lý sự kiện khi tất cả các tác vụ tìm kiếm hoàn tất (hoặc bị dừng)."""
        found_count = sum(1 for r in self.results_list if r.found == "Có")
        completion_msg = f"{final_message} Tìm thấy {found_count} kết quả khớp."

        if self.cb_auto_xlsx.isChecked() and self.results_list:
            self.update_status("Đang tự động lưu kết quả ra Excel...")
            self.auto_save_excel_file() # Đổi tên hàm cho rõ ràng

        self.cleanup_after_search_processing(completion_msg, True)

    def cleanup_after_search_processing(self, final_status_message, show_completion_messagebox):
        """Dọn dẹp tài nguyên và cập nhật UI sau khi tìm kiếm kết thúc."""
        self.elapsed_timer_display.stop()
        self.update_progress_bar(100) # Đặt tiến trình là 100%
        QTimer.singleShot(2000, lambda: self.progress_bar.setVisible(False)) # Ẩn progress bar sau 2s

        self.btn_search.setEnabled(True)
        self.btn_stop_search.setEnabled(False) # Đảm bảo nút stop bị vô hiệu hóa

        # Việc dọn dẹp QThread và SearchWorker đã được cấu hình để tự động thông qua
        # self.search_thread.finished.connect(self.search_thread.deleteLater)
        # self.search_worker_obj.destroyed.connect(self._on_worker_destroyed)
        # QThread sẽ phát tín hiệu finished khi hàm run() của worker kết thúc.
        
        self.update_status(final_status_message)
        if show_completion_messagebox:
            QMessageBox.information(self, "Hoàn tất Tìm kiếm", final_status_message)
        
        self.active_web_callbacks.clear() # Xóa các callbacks không còn dùng
        self.current_search_tasks_list.clear()
        # Không đặt self.search_thread = None ở đây, vì deleteLater sẽ xử lý nó.
        # self.search_worker_obj sẽ được đặt thành None trong _on_worker_destroyed.

    def add_result_to_grid_display(self, website, keyword, link, found_status, full_data):
        """Thêm một kết quả tìm kiếm vào bảng và danh sách results_list."""
        summary_length = 300 # Tăng độ dài tóm tắt
        summary = (full_data[:summary_length] + "...") if len(full_data) > summary_length else full_data
        result = SearchResult(website, keyword, link, found_status, full_data, summary)
        self.results_list.append(result)

        row_position = self.dgv_results.rowCount()
        self.dgv_results.insertRow(row_position)
        
        self.dgv_results.setItem(row_position, 0, QTableWidgetItem(website))
        self.dgv_results.setItem(row_position, 1, QTableWidgetItem(keyword))
        item_link = QTableWidgetItem(link)
        item_link.setToolTip(link) # Hiển thị link đầy đủ khi hover
        self.dgv_results.setItem(row_position, 2, item_link)
        self.dgv_results.setItem(row_position, 3, QTableWidgetItem(found_status))
        self.dgv_results.setItem(row_position, 4, QTableWidgetItem(summary))

        self.dgv_results.resizeRowsToContents() # Điều chỉnh chiều cao hàng cho vừa nội dung
        self.update_status(f"Đã thêm: {website} - '{keyword}' ({found_status})")

    def open_link_from_table(self, row, column):
        """Mở link trong QWebEngineView khi double click vào ô link trên bảng."""
        if column == 2: # Cột "Link"
            link_item = self.dgv_results.item(row, column)
            if link_item and link_item.text():
                url_str = link_item.text()
                if url_str.startswith("http://") or url_str.startswith("https://"):
                    self.web_view.load(QUrl(url_str))
                    self.update_status(f"Đang tải link: {url_str[:80]}...")
                else:
                    self.update_status(f"Link không hợp lệ: {url_str}")


    def auto_save_excel_file(self):
        """Tự động lưu kết quả tìm kiếm ra file Excel."""
        if not self.results_list:
            self.update_status("Không có kết quả để lưu ra Excel.")
            return
        try:
            # Tạo thư mục lưu trữ trên Desktop nếu chưa có
            desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
            folder_name = "IM_Keyword_Reports_Py" # Tên thư mục
            folder_path = os.path.join(desktop_path, folder_name)
            os.makedirs(folder_path, exist_ok=True) # Tạo thư mục, không báo lỗi nếu đã tồn tại

            base_filename = "KeywordSearchResults"
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            file_path = os.path.join(folder_path, f"{base_filename}_{timestamp}.xlsx")

            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "SearchResults"

            headers = ["Website", "Keyword", "Link", "Found", "Full Data Summary"]
            sheet.append(headers)
            # Định dạng đậm cho dòng tiêu đề
            for cell in sheet[1]: cell.font = Font(bold=True)

            for r_idx, r_data in enumerate(self.results_list, start=2): # Bắt đầu từ hàng 2
                sheet.cell(row=r_idx, column=1, value=r_data.website)
                sheet.cell(row=r_idx, column=2, value=r_data.keyword)
                link_cell_val = r_data.link
                link_cell = sheet.cell(row=r_idx, column=3, value=link_cell_val)
                # Tạo hyperlink nếu là URL hợp lệ
                if link_cell_val and (link_cell_val.startswith("http://") or link_cell_val.startswith("https://")):
                    try:
                        link_cell.hyperlink = link_cell_val
                        link_cell.style = "Hyperlink" # Áp dụng style hyperlink của Excel
                    except Exception as e_hyper:
                        if DEBUG_MAIN_THREAD_HANDLERS: print(f"MAIN: Error setting hyperlink for {link_cell_val}: {e_hyper}")
                sheet.cell(row=r_idx, column=4, value=r_data.found)
                sheet.cell(row=r_idx, column=5, value=r_data.full_data_summary) # Lưu tóm tắt
            
            # Tự động điều chỉnh độ rộng cột
            column_letters = ['A', 'B', 'C', 'D', 'E']
            for i, column_letter in enumerate(column_letters):
                max_len = 0
                for row_idx in range(1, sheet.max_row + 1):
                    cell_val = sheet.cell(row=row_idx, column=i+1).value
                    if cell_val:
                        max_len = max(max_len, len(str(cell_val)))
                
                # Giới hạn chiều rộng tối đa, đặc biệt cho cột link và summary
                adjusted_width = min(max_len + 5, 75 if column_letter in ['C', 'E'] else 45)
                sheet.column_dimensions[column_letter].width = adjusted_width

            workbook.save(file_path)
            self.update_status(f"File Excel đã lưu: {os.path.basename(file_path)}")
            # Không hiển thị QMessageBox ở đây nữa nếu là auto_save
            # QMessageBox.information(self, "Xuất Excel Thành công", f"Kết quả đã được lưu vào:\n{file_path}")

        except Exception as e:
            self.update_status(f"Lỗi khi lưu file Excel: {e}")
            QMessageBox.critical(self, "Lỗi Xuất Excel", f"Không thể lưu file Excel: {e}\nChi tiết: {type(e)}")

    def export_to_excel(self):
        """Xử lý sự kiện click nút Xuất Excel (gọi hàm auto_save và hiển thị thông báo)."""
        if not self.results_list:
            QMessageBox.information(self, "Không có Dữ liệu", "Không có kết quả để xuất ra Excel.")
            return
        self.auto_save_excel_file() # Gọi hàm lưu tự động
        # Tìm file mới nhất đã lưu để hiển thị đường dẫn cho người dùng
        desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
        folder_name = "IM_Keyword_Reports_Py"
        folder_path = os.path.join(desktop_path, folder_name)
        try:
            files = [os.path.join(folder_path, f) for f in os.listdir(folder_path) if f.startswith("KeywordSearchResults") and f.endswith(".xlsx")]
            if files:
                latest_file = max(files, key=os.path.getctime)
                QMessageBox.information(self, "Xuất Excel Thành công", f"Kết quả đã được lưu vào:\n{latest_file}")
            else: # Trường hợp auto_save_excel_file bị lỗi và không tạo file
                 QMessageBox.warning(self, "Xuất Excel", "Đã cố gắng lưu nhưng có thể đã xảy ra lỗi. Vui lòng kiểm tra log.")
        except Exception: # Nếu có lỗi khi tìm file
            QMessageBox.information(self, "Xuất Excel", "Đã cố gắng lưu file Excel. Vui lòng kiểm tra thư mục 'IM_Keyword_Reports_Py' trên Desktop.")


    def show_login_menu(self):
        """Hiển thị menu để người dùng chọn nền tảng và điều hướng đến trang đăng nhập."""
        menu = QMenu(self)
        actions_data = [
            ("Đăng nhập Instagram", "https://www.instagram.com/accounts/login/"),
            ("Đăng nhập Facebook", "https://www.facebook.com/login/"),
            ("Đăng nhập X (Twitter)", "https://x.com/login")
        ]
        for text, url_str in actions_data:
            action = QAction(text, self)
            # Sử dụng lambda để truyền URL vào slot
            action.triggered.connect(lambda checked, u=url_str: self.web_view.load(QUrl(u)))
            menu.addAction(action)
        # Hiển thị menu tại vị trí của nút login
        menu.popup(self.btn_login.mapToGlobal(self.btn_login.rect().bottomLeft()))

    def clear_browser_cache(self):
        """Xóa cache HTTP và cookies của QWebEngineView."""
        if self.web_view:
            try:
                self.update_status("Đang xóa cache và cookies của WebView...")
                profile = self.web_view.page().profile()
                profile.clearHttpCache() # Xóa cache HTTP
                profile.cookieStore().deleteAllCookies() # Xóa tất cả cookies
                # profile.clearAllVisitedLinks() # Tùy chọn: xóa cả lịch sử duyệt web trong session này
                QMessageBox.information(self, "Đã xóa Cache", "Cache HTTP và Cookies của WebView đã được xóa thành công.")
                self.update_status("Cache và Cookies của WebView đã được xóa.")
                self.web_view.reload() # Tải lại trang hiện tại để thấy hiệu quả
            except Exception as e:
                QMessageBox.critical(self, "Lỗi Xóa Cache", f"Đã xảy ra lỗi khi xóa cache: {e}")
        else:
            QMessageBox.warning(self, "Lỗi WebView", "Đối tượng WebView chưa được khởi tạo.")

    def scroll_to_keyword_on_main(self, keyword):
        """Cuộn QWebEngineView đến vị trí chứa từ khóa (nếu tìm thấy)."""
        self.update_status(f"Đang thử cuộn đến từ khóa '{keyword}'...")
        try:
            # Đảm bảo từ khóa được escape đúng cách để dùng trong chuỗi JavaScript
            sanitized_keyword_for_js = json.dumps(keyword.lower())[1:-1] # Lấy phần bên trong dấu ""
            js_scroll_script = f"""
                (function() {{
                    var keywordToFind = '{sanitized_keyword_for_js}';
                    var elements = Array.from(document.querySelectorAll('body *:not(script):not(style):not(path):not(svg)'));
                    var foundElement = null;
                    
                    // Tìm phần tử chứa text khớp nhất và có thể nhìn thấy
                    for (let el of elements) {{
                        if (el.innerText && el.innerText.toLowerCase().includes(keywordToFind)) {{
                            // Ưu tiên các phần tử không phải là link hoặc button nếu có thể
                            if (el.offsetWidth > 0 && el.offsetHeight > 0) {{ // Kiểm tra xem có hiển thị không
                                foundElement = el;
                                if (el.tagName !== 'A' && el.tagName !== 'BUTTON') break; 
                            }}
                        }}
                    }}

                    if (foundElement) {{
                        foundElement.scrollIntoView({{behavior: 'smooth', block: 'center', inline: 'nearest'}});
                        // Tùy chọn: thêm highlight tạm thời để người dùng dễ thấy
                        // foundElement.style.outline = '3px solid red';
                        // setTimeout(() => {{ foundElement.style.outline = ''; }}, 3500);
                        return true;
                    }}
                    return false;
                }})();"""
            self.web_view.page().runJavaScript(js_scroll_script, 
                lambda res: self.update_status(f"Kết quả cuộn đến '{keyword}': {'Thành công' if res else 'Không tìm thấy phần tử'}"))
        except Exception as e:
            self.update_status(f"Lỗi khi cuộn đến từ khóa: {e}")

    def capture_screenshot_on_main(self, website, keyword, identifier_url_or_ts):
        """Chụp ảnh màn hình của QWebEngineView."""
        self.update_status(f"Đang chuẩn bị chụp ảnh màn hình cho {website} - '{keyword}'...")
        try:
            desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
            folder_name = "IM_Keyword_Screenshots_Py"
            folder_path = os.path.join(desktop_path, folder_name)
            os.makedirs(folder_path, exist_ok=True)

            # Tạo tên file an toàn và dễ nhận biết
            safe_website = "".join(c if c.isalnum() else "_" for c in website)
            safe_keyword = "".join(c if c.isalnum() else "_" for c in keyword)
            
            # Làm sạch identifier (có thể là URL hoặc timestamp)
            if "http" in identifier_url_or_ts: # Nếu là URL
                # Lấy phần cuối của URL làm định danh, giới hạn độ dài
                safe_identifier = identifier_url_or_ts.split('/')[-1] if identifier_url_or_ts.split('/')[-1] else identifier_url_or_ts.split('/')[-2]
                safe_identifier = "".join(c if c.isalnum() else "_" for c in safe_identifier)[:35]
            else: # Nếu là timestamp hoặc định danh khác
                safe_identifier = str(identifier_url_or_ts)

            timestamp_now = datetime.datetime.now().strftime("%H%M%S")
            file_name = f"{safe_website}_{safe_keyword}_{safe_identifier}_{timestamp_now}.png"
            file_path = os.path.join(folder_path, file_name)

            if self.cb_long_screenshot.isChecked():
                self.update_status(f"Đang thử chụp ảnh dài cho '{keyword}'...")
                # Lấy chiều cao thực của tài liệu HTML
                js_get_height = "Math.max( document.body.scrollHeight, document.body.offsetHeight, document.documentElement.clientHeight, document.documentElement.scrollHeight, document.documentElement.offsetHeight );"
                self.web_view.page().runJavaScript(js_get_height,
                    lambda height_js_result: self._handle_long_screenshot_capture(height_js_result, file_path, website, keyword))
            else: # Chụp ảnh màn hình thường (phần nhìn thấy)
                self._save_normal_screenshot(file_path, website, keyword)
        except Exception as e:
            self.update_status(f"Lỗi nghiêm trọng khi chuẩn bị chụp ảnh màn hình: {e}")

    def _handle_long_screenshot_capture(self, height_js_result, file_path, website, keyword):
        """Xử lý logic chụp ảnh màn hình dài sau khi có chiều cao từ JS."""
        original_size = self.web_view.size() # Lưu kích thước gốc của webview
        try:
            doc_height_str = str(height_js_result) # Kết quả từ JS có thể là số hoặc chuỗi
            
            if doc_height_str and doc_height_str.replace('.', '', 1).isdigit(): # Kiểm tra xem có phải là số không
                doc_height = int(float(doc_height_str)) # Chuyển sang float rồi int để xử lý số thập phân
                view_width = original_size.width()
                
                # Giới hạn chiều cao tối đa để tránh lỗi bộ nhớ hoặc treo (ví dụ: 16384 pixels)
                max_screenshot_height = 16384 

                if doc_height > original_size.height() and doc_height > 0:
                    capture_height = min(doc_height, max_screenshot_height)
                    self.update_status(f"Đang thay đổi kích thước WebView để chụp ảnh dài (cao: {capture_height}px)...")
                    self.web_view.resize(view_width, capture_height)
                    # Phải chờ một chút để webview render lại với kích thước mới trước khi chụp
                    # Sử dụng QTimer để đảm bảo việc này xảy ra trong event loop của Qt
                    QTimer.singleShot(1500, lambda: self._save_resized_screenshot_and_restore(file_path, original_size, website, keyword))
                    return # Việc lưu sẽ được thực hiện trong _save_resized_screenshot_and_restore
                else: 
                    self.update_status(f"Chiều cao tài liệu ({doc_height}px) không lớn hơn view hiện tại. Chụp ảnh thường.")
                    self._save_normal_screenshot(file_path, website, keyword)
            else:
                self.update_status(f"Không lấy được chiều cao tài liệu hợp lệ từ JS ('{height_js_result}'). Chụp ảnh thường.")
                self._save_normal_screenshot(file_path, website, keyword)

        except Exception as e:
            self.update_status(f"Lỗi trong quá trình xử lý chụp ảnh dài: {e}")
            self._save_normal_screenshot(file_path, website, keyword) # Fallback chụp thường
            self.web_view.resize(original_size) # Đảm bảo khôi phục kích thước nếu có lỗi ở đây

    def _save_resized_screenshot_and_restore(self, file_path, original_size, website, keyword):
        """Lưu ảnh sau khi webview đã được resize và sau đó khôi phục kích thước gốc."""
        try:
            # Tạo QPixmap với kích thước hiện tại của webview (đã được resize)
            pixmap = QPixmap(self.web_view.size()) 
            self.web_view.render(pixmap) # Render nội dung webview vào pixmap
            
            if pixmap.save(file_path, "PNG"):
                self.update_status(f"Ảnh màn hình dài đã lưu: {os.path.basename(file_path)}")
            else:
                self.update_status(f"Lỗi: Không thể lưu ảnh màn hình dài: {os.path.basename(file_path)}")
        except Exception as e:
             self.update_status(f"Lỗi khi lưu ảnh dài đã resize: {e}")
        finally:
            self.web_view.resize(original_size) # Luôn khôi phục kích thước gốc của webview

    def _save_normal_screenshot(self, file_path, website, keyword):
        """Lưu ảnh màn hình thường (phần đang hiển thị)."""
        try:
            pixmap = self.web_view.grab() # Chụp phần đang hiển thị
            if pixmap.save(file_path, "PNG"):
                self.update_status(f"Ảnh màn hình (thường/fallback) đã lưu: {os.path.basename(file_path)}")
            else:
                self.update_status(f"Lỗi: Không thể lưu ảnh (thường/fallback): {os.path.basename(file_path)}")
        except Exception as e:
            self.update_status(f"Lỗi khi lưu ảnh thường/fallback: {e}")

    def closeEvent(self, event):
        """Xử lý sự kiện đóng cửa sổ ứng dụng."""
        if self.search_thread and self.search_thread.isRunning():
            reply = QMessageBox.question(self, 'Đang thoát Ứng dụng...',
                                         "Một quá trình tìm kiếm đang chạy. Bạn có chắc chắn muốn dừng và thoát không?",
                                         QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
            if reply == QMessageBox.Yes:
                self.update_status("Đang dừng các tác vụ và chuẩn bị thoát...")
                if self.search_worker_obj:
                    self.search_worker_obj.stop() # Yêu cầu worker dừng
                
                # Worker nên tự kết thúc. QThread.finished sẽ được phát ra.
                # Tuy nhiên, vì ứng dụng đang đóng, chúng ta không muốn đợi quá lâu.
                # Cho worker một khoảng thời gian ngắn để dọn dẹp.
                if self.search_thread.isRunning():
                    if DEBUG_MAIN_THREAD_HANDLERS: print("MAIN (CloseEvent): Requesting thread quit and waiting briefly.")
                    self.search_thread.quit() # Yêu cầu QThread thoát vòng lặp sự kiện của nó
                    if not self.search_thread.wait(2000): # Chờ tối đa 2 giây
                        if DEBUG_MAIN_THREAD_HANDLERS: print("MAIN (CloseEvent): Thread did not finish in time. App will close anyway.")
                        # Không nên gọi terminate() vì có thể gây rò rỉ tài nguyên hoặc crash.
                        # Hệ điều hành sẽ xử lý khi chương trình chính đóng.
                event.accept() # Chấp nhận sự kiện đóng
            else:
                event.ignore() # Bỏ qua sự kiện đóng, ứng dụng tiếp tục chạy
        else:
            event.accept() # Không có tìm kiếm nào đang chạy, đóng bình thường


if __name__ == "__main__":
    app = QApplication(sys.argv)
    
    # Tùy chọn: Cấu hình profile mặc định của QWebEngine (nếu cần)
    # QWebEngineProfile.defaultProfile().setPersistentCookiesPolicy(QWebEngineProfile.AllowPersistentCookies) # Cho phép lưu cookies lâu dài
    # QWebEngineProfile.defaultProfile().setHttpCacheType(QWebEngineProfile.DiskHttpCache) # Sử dụng cache đĩa (có thể tăng tốc tải trang lặp lại)
    # os.environ["QTWEBENGINE_REMOTE_DEBUGGING"] = "9223" # Bật remote debugging (truy cập localhost:9223 từ Chrome/Edge)

    main_win = KeywordSearchApp()
    main_win.show()
    sys.exit(app.exec_())
